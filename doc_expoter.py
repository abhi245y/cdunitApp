import pymongo
import pandas as pd
from bson.int64 import Int64
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

# Connect to MongoDB
client = pymongo.MongoClient("mongodb://10.20.80.108:27017/")
db = client["cd_unit"]
collection = db["bundleDetails"]

# Function to retrieve documents within given qpCode range and series
def get_documents_in_range(start_series, end_series, start_code, end_code):
    query = {
        "$and": [
            {"qpSeries": {"$gte": start_series, "$lte": end_series}},
            {"$or": [
                {"qpCode": {"$gte": start_code, "$lte": end_code}},
                {"qpCode": {"$regex": "^[0-9]+$", "$gte": start_code, "$lte": end_code}}
            ]}
        ]
    }
    documents = list(collection.find(query))
    return documents

# Function to create Excel table from documents
def create_excel_table(documents):
    df = pd.DataFrame(documents)
    df.drop("_id", axis=1, inplace=True)  # Drop MongoDB _id field
    
    # Define formatting options
    format_mapping = {
        "receivedDate": lambda x: pd.to_datetime(x, unit="ms").strftime("%Y-%m-%d"), # Convert receivedDate to readable date format
        "isNil": lambda x: "Yes" if x else "No"  # Convert isNil to Yes/No
    }
    
    # Apply formatting
    for col, formatter in format_mapping.items():
        if col in df.columns:
            df[col] = df[col].apply(formatter)
    
    # Write to Excel
    writer = pd.ExcelWriter("bundle_details_formatted.xlsx", engine='openpyxl')
    df.to_excel(writer, index=False)
    worksheet = writer.sheets['Sheet1']
    
    # Apply additional formatting options
    for idx, col in enumerate(df): 
        max_len = df[col].astype(str).str.len().max()
        max_len = max_len if len(col) <= max_len else len(col)
        worksheet.column_dimensions[chr(65 + idx)].width = max_len + 2  # Adjust column width to fit content
        worksheet.cell(row=1, column=idx+1).font = worksheet.cell(row=1, column=idx+1).font.copy(bold=True)  # Bold column headers
        worksheet.cell(row=1, column=idx+1).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))  # Add border to column headers

    # Create Excel table
    table = Table(displayName="Table1", ref=worksheet.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    worksheet.add_table(table)
    
    writer.save()

# Main function
def main():
    start_series = "S"
    end_series = "S"
    start_code = "1202"
    end_code = "1203"

    documents = get_documents_in_range(start_series, end_series, start_code, end_code)
    create_excel_table(documents)

if __name__ == "__main__":
    main()
